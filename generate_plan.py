from __future__ import annotations

import io
from copy import copy

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from planner import FOCUS_TEXT, GOAL_TEXT_BY_MACRO, PHASE_COLORS, build_plan_context


HEADER_BG = (64, 64, 64)
WHITE = (255, 255, 255)
BLACK = (0, 0, 0)
LIGHT_GREY = (217, 217, 217)


def rgb_hex(rgb: tuple[int, int, int]) -> str:
    return "".join(f"{value:02X}" for value in rgb)


def tinted(rgb: tuple[int, int, int], ratio: float = 0.5) -> tuple[int, int, int]:
    return tuple(int(value + (255 - value) * ratio) for value in rgb)


def font_color_for_fill(rgb: tuple[int, int, int]) -> str:
    luminance = (0.299 * rgb[0]) + (0.587 * rgb[1]) + (0.114 * rgb[2])
    return rgb_hex(WHITE if luminance < 140 else BLACK)


def apply_fill(cell, rgb: tuple[int, int, int]) -> None:
    cell.fill = PatternFill(fill_type="solid", fgColor=rgb_hex(rgb))


def apply_border(cell, border: Border) -> None:
    cell.border = copy(border)


def add_all_borders(ws, max_col: int, max_row: int, base_border: Border) -> None:
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            apply_border(cell, base_border)


def add_medium_boundary(ws, row: int, column: int) -> None:
    cell = ws.cell(row=row, column=column)
    cell.border = Border(
        left=cell.border.left,
        right=Side(style="medium", color=rgb_hex(BLACK)),
        top=cell.border.top,
        bottom=cell.border.bottom,
    )


def stack_rows_for_level(level: int) -> list[int]:
    bounded = max(0, min(5, level))
    return list(range(19 - bounded, 19)) if bounded else []


def merge_and_style(ws, start_row: int, start_col: int, end_row: int, end_col: int, value: str, *, fill=None, font=None, alignment=None):
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    cell = ws.cell(row=start_row, column=start_col)
    cell.value = value
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment
    return cell


def create_season_plan(data: dict) -> bytes:
    context = build_plan_context(data)
    year = context["athlete"]["year"]
    rider_first = context["athlete"]["riderFirst"]
    rider_last = context["athlete"]["riderLast"]
    coach_first = context["athlete"]["coachFirst"]
    coach_last = context["athlete"]["coachLast"]
    week_starts = context["weekStarts"]
    plan_weeks = context["planWeeks"]
    phase_blocks = context["phaseBlocks"]
    macro_blocks = context["macroBlocks"]
    total_weeks = context["totalWeeks"]
    race_cells = context["raceCells"]

    wb = Workbook()
    ws = wb.active
    ws.title = f"Season Plan {year}"
    ws.freeze_panes = "B4"

    thin_side = Side(style="thin", color=rgb_hex(BLACK))
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    dark_fill = PatternFill(fill_type="solid", fgColor=rgb_hex(HEADER_BG))
    light_grey_fill = PatternFill(fill_type="solid", fgColor=rgb_hex(LIGHT_GREY))

    max_col = total_weeks + 1
    ws.column_dimensions["A"].width = 22
    for column in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(column)].width = 4.5

    row_heights = {
        1: 30,
        2: 20,
        4: 15,
        5: 15,
        6: 15,
        7: 90,
        8: 15,
        9: 15,
        10: 25,
        11: 30,
        12: 40,
        13: 65,
        14: 14,
        15: 14,
        16: 14,
        17: 14,
        18: 14,
        19: 130,
        20: 110,
    }
    for row_number, height in row_heights.items():
        ws.row_dimensions[row_number].height = height

    header_end_col = min(8, max_col)
    merge_and_style(
        ws,
        1,
        1,
        1,
        header_end_col,
        "Season Plan",
        fill=dark_fill,
        font=Font(name="Arial", size=16, bold=True, color=rgb_hex(WHITE)),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    if max_col > header_end_col:
        merge_and_style(
            ws,
            1,
            header_end_col + 1,
            1,
            max_col,
            f"Last name, Name: {coach_last} {coach_first}",
            fill=dark_fill,
            font=Font(name="Arial", size=11, bold=True, color=rgb_hex(WHITE)),
            alignment=Alignment(horizontal="right", vertical="center"),
        )

    if max_col >= 2:
        rider_split = max(2, min(max_col, 6))
        merge_and_style(
            ws,
            2,
            1,
            2,
            rider_split,
            f"Rider: {rider_first} {rider_last}",
            font=Font(name="Arial", size=11, bold=True),
            alignment=Alignment(horizontal="left", vertical="center"),
        )
        if rider_split < max_col:
            merge_and_style(
                ws,
                2,
                rider_split + 1,
                2,
                max_col,
                f"Year: {year}",
                font=Font(name="Arial", size=11, bold=True),
                alignment=Alignment(horizontal="right", vertical="center"),
            )

    ws["A4"] = "Month"
    ws["A5"] = "Week Commencing"
    ws["A6"] = ""
    ws["A7"] = "Competition (Identify target event)"
    ws["A8"] = "Testing FTP"
    ws["A9"] = "4D Profile Testing (no20')"
    ws["A10"] = "Macro Cycle"
    ws["A11"] = "(Training Phases and sub-phases)"
    ws["A12"] = "Meso Cycle"
    ws["A13"] = "Micro Cycle"
    ws["A14"] = "5 (high)"
    ws["A15"] = "4"
    ws["A16"] = "3"
    ws["A17"] = "2"
    ws["A18"] = "1 (low)"
    ws["A19"] = "Training Focus"
    ws["A20"] = "Goals"

    for label_row in range(4, 21):
        ws[f"A{label_row}"].font = Font(name="Arial", size=10, bold=True)
        ws[f"A{label_row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    month_runs = []
    if week_starts:
        from datetime import date

        parsed_mondays = [date.fromisoformat(item) for item in week_starts]
        start_index = 1
        current_key = (parsed_mondays[0].year, parsed_mondays[0].month)
        for index, monday in enumerate(parsed_mondays[1:], start=2):
            key = (monday.year, monday.month)
            if key != current_key:
                month_runs.append((start_index, index - 1, parsed_mondays[start_index - 1].strftime("%B")))
                start_index = index
                current_key = key
        month_runs.append((start_index, len(parsed_mondays), parsed_mondays[start_index - 1].strftime("%B")))

    for start_week, end_week, month_name in month_runs:
        merge_and_style(
            ws,
            4,
            start_week + 1,
            4,
            end_week + 1,
            month_name,
            font=Font(name="Arial", size=10, bold=True),
            alignment=Alignment(horizontal="center", vertical="center"),
        )

    for week in plan_weeks:
        column = week["week"] + 1
        ws.cell(row=5, column=column).value = week["week"]
        ws.cell(row=6, column=column).value = week["day"]
        ws.cell(row=5, column=column).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=6, column=column).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=5, column=column).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=6, column=column).font = Font(name="Arial", size=10)

    for race_week, payload in race_cells.items():
        target_cell = ws.cell(row=7, column=race_week + 1)
        event_names = []
        for event in payload.get("events", []):
            prefix = ""
            if event["type"] == "training_camp":
                prefix = "Camp: "
            event_names.append(f"{prefix}{event['name']}")
        target_cell.value = "\n".join(event_names or payload["names"])
        target_cell.alignment = Alignment(textRotation=90, horizontal="center", vertical="center", wrap_text=True)
        if payload.get("has_main"):
            fill_color = PHASE_COLORS["Race"]
            font_color = rgb_hex(WHITE)
        elif payload.get("has_secondary_race"):
            fill_color = PHASE_COLORS["Taper"]
            font_color = rgb_hex(BLACK)
        else:
            fill_color = PHASE_COLORS["Recovery"]
            font_color = rgb_hex(BLACK)
        target_cell.font = Font(name="Arial", size=9, bold=True, color=font_color)
        apply_fill(target_cell, fill_color)

    ftp_test_weeks = {week["week"] for week in plan_weeks if week.get("ftpTest")}
    profile_test_weeks = {week["week"] for week in plan_weeks if week.get("fourDTest")}

    for week_number in ftp_test_weeks:
        cell = ws.cell(row=8, column=week_number + 1)
        cell.value = "T"
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        apply_fill(cell, PHASE_COLORS["Build 1"])

    for week_number in profile_test_weeks:
        cell = ws.cell(row=9, column=week_number + 1)
        cell.value = "T"
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        apply_fill(cell, PHASE_COLORS["Build 1"])

    if total_weeks >= 1:
        merge_and_style(
            ws,
            10,
            2,
            10,
            max_col,
            f"Season {year} A",
            fill=PatternFill(fill_type="solid", fgColor=rgb_hex(tinted(HEADER_BG, 0.15))),
            font=Font(name="Arial", size=14, bold=True),
            alignment=Alignment(horizontal="center", vertical="center"),
        )

    for block in macro_blocks:
        merge_and_style(
            ws,
            11,
            block["startWeek"] + 1,
            11,
            min(block["endWeek"] + 1, max_col),
            block["macro"],
            fill=light_grey_fill,
            font=Font(name="Arial", size=10, bold=True),
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        )

    row_12_titles = {
        "Prep": "PREP",
        "Base 1": "BASE 1",
        "Base 2": "BASE 2",
        "Base 3": "BASE 3",
        "Build 1": "BUILD 1",
        "Build 2": "BUILD 2",
        "Peak": "PEAK",
        "Taper": "TAPER",
        "Race": "RACE",
        "Recovery": "RECOVERY",
    }

    for block in phase_blocks:
        phase_name = block["phase"]
        start_week = block["startWeek"]
        end_week = block["endWeek"]
        color = PHASE_COLORS[phase_name]
        merge_and_style(
            ws,
            12,
            start_week + 1,
            12,
            end_week + 1,
            row_12_titles[phase_name],
            fill=PatternFill(fill_type="solid", fgColor=rgb_hex(color)),
            font=Font(name="Arial", size=12, bold=True, color=font_color_for_fill(color)),
            alignment=Alignment(horizontal="center", vertical="center"),
        )

    for week in plan_weeks:
        column = week["week"] + 1
        color = PHASE_COLORS[week["phase"]]
        cell = ws.cell(row=13, column=column)
        cell.value = week["micro"]
        cell.font = Font(name="Arial", size=8, bold=True, color=font_color_for_fill(color))
        cell.alignment = Alignment(horizontal="center", vertical="center", textRotation=90, wrap_text=True)
        apply_fill(cell, color)

        for row_number in stack_rows_for_level(week["volume"]):
            apply_fill(ws.cell(row=row_number, column=column), tinted(color, 0.2))
        for row_number in stack_rows_for_level(week["intensity"]):
            marker = ws.cell(row=row_number, column=column)
            marker.value = "X"
            marker.font = Font(name="Arial", size=8, bold=True, color=rgb_hex(BLACK))
            marker.alignment = Alignment(horizontal="center", vertical="center")

    focus_runs = []
    if plan_weeks:
        run_start = 1
        current_phase = plan_weeks[0]["phase"]
        current_text = plan_weeks[0]["focusText"]
        for week in plan_weeks[1:]:
            if week["phase"] != current_phase or week["focusText"] != current_text:
                focus_runs.append((run_start, week["week"] - 1, current_phase, current_text))
                run_start = week["week"]
                current_phase = week["phase"]
                current_text = week["focusText"]
        focus_runs.append((run_start, plan_weeks[-1]["week"], current_phase, current_text))

    for start_week, end_week, phase_name, focus_text in focus_runs:
        merge_and_style(
            ws,
            19,
            start_week + 1,
            19,
            end_week + 1,
            focus_text,
            fill=PatternFill(fill_type="solid", fgColor=rgb_hex(tinted(PHASE_COLORS[phase_name], 0.5))),
            font=Font(name="Arial", size=8),
            alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
        )

    macro_fill_map = {
        "General Preparation": tinted(PHASE_COLORS["Prep"], 0.6),
        "Pre-Competition": tinted(PHASE_COLORS["Build 1"], 0.6),
        "Competition": tinted(PHASE_COLORS["Peak"], 0.6),
        "Transition": tinted(PHASE_COLORS["Recovery"], 0.6),
    }

    goal_runs = []
    if plan_weeks:
        run_start = 1
        current_macro = plan_weeks[0]["macro"]
        current_text = plan_weeks[0]["goalText"]
        for week in plan_weeks[1:]:
            if week["macro"] != current_macro or week["goalText"] != current_text:
                goal_runs.append((run_start, week["week"] - 1, current_macro, current_text))
                run_start = week["week"]
                current_macro = week["macro"]
                current_text = week["goalText"]
        goal_runs.append((run_start, plan_weeks[-1]["week"], current_macro, current_text))

    for start_week, end_week, macro_name, goal_text in goal_runs:
        merge_and_style(
            ws,
            20,
            start_week + 1,
            20,
            end_week + 1,
            goal_text,
            fill=PatternFill(fill_type="solid", fgColor=rgb_hex(macro_fill_map[macro_name])),
            font=Font(name="Arial", size=8),
            alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
        )

    add_all_borders(ws, max_col, 20, thin_border)
    for block in phase_blocks:
        add_medium_boundary(ws, 12, block["endWeek"] + 1)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = f"A1:{get_column_letter(max_col)}20"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
